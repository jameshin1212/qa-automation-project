"""
Generate Excel test case documentation using openpyxl only
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

def create_test_cases_excel():
    """Create comprehensive test case documentation in Excel"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Category", "API Tests", "UI Tests", "Total", "Automated", "Manual", "Coverage"])
    ws_summary.append(["정상 플로우", 4, 1, 5, 5, 0, "100%"])
    ws_summary.append(["비정상 플로우", 8, 3, 11, 11, 0, "100%"])
    ws_summary.append(["경계값", 5, 0, 5, 5, 0, "100%"])
    ws_summary.append(["보안", 6, 0, 6, 6, 0, "100%"])
    ws_summary.append(["중복 방지", 2, 1, 3, 3, 0, "100%"])
    ws_summary.append(["UI", 0, 1, 1, 1, 0, "100%"])
    ws_summary.append(["합계", 25, 6, 31, 31, 0, "100%"])
    
    # Style the header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Create Test Cases sheet
    ws_cases = wb.create_sheet("Test Cases")
    headers = ["TC_ID", "Category", "Test_Name", "Test_Type", "Priority", "Test_Method",
               "Precondition", "Test_Steps", "Test_Data", "Expected_Result", 
               "Automation_Status", "Test_Script"]
    ws_cases.append(headers)
    
    # Add test cases
    test_cases = [
        ["TC-001", "정상 플로우", "정상적인 이메일과 비밀번호로 회원가입", "API", "Critical", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: valid.user@test.com\npassword: Test1234!", "201 Created, 사용자 생성됨", "완료",
         "test_registration_positive.py::test_registration_valid_email_password_success"],
        
        ["TC-002", "정상 플로우", "이메일에 + 기호가 포함된 경우", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. + 기호가 포함된 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: user+tag@gmail.com\npassword: Secure@Pass123", "201 Created, + 기호 허용", "완료",
         "test_registration_positive.py::test_registration_email_with_plus_sign_success"],
        
        ["TC-003", "정상 플로우", "한국 도메인 이메일로 회원가입", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. naver.com 도메인 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: korean.user@naver.com\npassword: Korean123!@#", "201 Created, 한국 도메인 허용", "완료",
         "test_registration_positive.py::test_registration_korean_domain_email_success"],
        
        ["TC-004", "정상 플로우", "긴 이메일 주소로 회원가입", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. 30자 이상 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: long.email@example.com", "201 Created, 긴 이메일 허용", "완료",
         "test_registration_positive.py::test_registration_long_email_address_success"],
        
        ["TC-005", "비정상 플로우", "잘못된 이메일 형식 - @ 기호 누락", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. @ 없는 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: invalid-email\npassword: Test1234!", "400 Bad Request, 에러 메시지", "완료",
         "test_registration_negative.py::test_registration_invalid_email_missing_at_fail"],
        
        ["TC-006", "비정상 플로우", "잘못된 이메일 - 로컬 파트 누락", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 로컬 파트 없는 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: @test.com\npassword: Test1234!", "400 Bad Request", "완료",
         "test_registration_negative.py::test_registration_invalid_email_missing_local_fail"],
        
        ["TC-007", "비정상 플로우", "잘못된 이메일 - 도메인 누락", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 도메인 없는 이메일 입력\n2. 유효한 비밀번호 입력\n3. 회원가입 요청",
         "email: test@\npassword: Test1234!", "400 Bad Request", "완료",
         "test_registration_negative.py::test_registration_invalid_email_missing_domain_fail"],
        
        ["TC-008", "비정상 플로우", "비밀번호가 너무 짧음", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일 입력\n2. 짧은 비밀번호 입력\n3. 회원가입 요청",
         "email: test@test.com\npassword: short", "400 Bad Request, 최소 8자 필요", "완료",
         "test_registration_negative.py::test_registration_password_too_short_fail"],
        
        ["TC-009", "비정상 플로우", "비밀번호에 대문자 누락", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일 입력\n2. 대문자 없는 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: nouppercase123!", "400 Bad Request, 대문자 필요", "완료",
         "test_registration_negative.py::test_registration_password_missing_uppercase_fail"],
        
        ["TC-010", "비정상 플로우", "비밀번호에 소문자 누락", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일 입력\n2. 소문자 없는 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: NOLOWERCASE123!", "400 Bad Request, 소문자 필요", "완료",
         "test_registration_negative.py::test_registration_password_missing_lowercase_fail"],
        
        ["TC-011", "비정상 플로우", "빈 이메일 필드", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 빈 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: (empty)\npassword: Test1234!", "400 Bad Request, 필수 필드", "완료",
         "test_registration_negative.py::test_registration_empty_email_fail"],
        
        ["TC-012", "비정상 플로우", "빈 비밀번호 필드", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일\n2. 빈 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: (empty)", "400 Bad Request, 필수 필드", "완료",
         "test_registration_negative.py::test_registration_empty_password_fail"],
        
        ["TC-013", "경계값", "최소 길이 이메일", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. 최소 길이 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: a@b.co\npassword: Pass123!", "201 Created", "완료",
         "test_registration_boundary.py::test_registration_minimum_email_length_success"],
        
        ["TC-014", "경계값", "최소 길이 비밀번호 (8자)", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일\n2. 정확히 8자 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: Pass123!", "201 Created", "완료",
         "test_registration_boundary.py::test_registration_minimum_password_length_success"],
        
        ["TC-015", "경계값", "매우 긴 이메일 주소", "API", "Low", "API_Auto",
         "Mock 서버 실행 중", "1. 100자 이상 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: very.long.email@test.com", "201 Created 또는 길이 제한", "완료",
         "test_registration_boundary.py::test_registration_very_long_email_success"],
        
        ["TC-016", "경계값", "매우 긴 비밀번호", "API", "Low", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일\n2. 100자 이상 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: VeryLongPassword...", "201 Created 또는 길이 제한", "완료",
         "test_registration_boundary.py::test_registration_very_long_password_success"],
        
        ["TC-017", "경계값", "이메일 공백 처리", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. 공백 포함 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: '  test@test.com  '\npassword: Test1234!", "공백 제거 후 처리", "완료",
         "test_registration_boundary.py::test_registration_email_with_spaces_trimmed"],
        
        ["TC-018", "보안", "SQL Injection - 이메일", "API", "Critical", "API_Auto",
         "Mock 서버 실행 중", "1. SQL Injection 시도\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: admin'--@test.com\npassword: Test1234!", "400 Bad Request, 차단됨", "완료",
         "test_registration_security.py::test_registration_sql_injection_email_blocked"],
        
        ["TC-019", "보안", "SQL Injection - 비밀번호", "API", "Critical", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일\n2. SQL Injection 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: ' OR '1'='1", "400 Bad Request, 차단됨", "완료",
         "test_registration_security.py::test_registration_sql_injection_password_blocked"],
        
        ["TC-020", "보안", "XSS 공격 - 이메일", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. XSS 스크립트 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: <script>alert('XSS')</script>@test.com", "400 Bad Request, 차단됨", "완료",
         "test_registration_security.py::test_registration_xss_email_sanitized"],
        
        ["TC-021", "보안", "XSS 공격 - 비밀번호", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 이메일\n2. XSS 스크립트 비밀번호\n3. 회원가입 요청",
         "email: test@test.com\npassword: <script>alert('XSS')</script>", "400 Bad Request, 차단됨", "완료",
         "test_registration_security.py::test_registration_xss_password_handled"],
        
        ["TC-022", "보안", "Path Traversal 시도", "API", "Medium", "API_Auto",
         "Mock 서버 실행 중", "1. Path traversal 이메일\n2. 유효한 비밀번호\n3. 회원가입 요청",
         "email: ../../../etc/passwd@test.com", "400 Bad Request, 차단됨", "완료",
         "test_registration_security.py::test_registration_path_traversal_email_blocked"],
        
        ["TC-023", "보안", "비밀번호 암호화 확인", "API", "Critical", "API_Auto",
         "Mock 서버 실행 중", "1. 유효한 회원가입\n2. 저장된 데이터 확인\n3. 비밀번호 평문 여부 검증",
         "email: security.test@example.com\npassword: SecurePass123!", "비밀번호 암호화됨", "완료",
         "test_registration_security.py::test_registration_password_not_stored_plain"],
        
        ["TC-024", "중복 방지", "중복 이메일 차단", "API", "Critical", "API_Auto",
         "Mock 서버 실행 중", "1. 첫 번째 회원가입\n2. 동일 이메일로 재시도\n3. 중복 확인",
         "email: duplicate@test.com\npassword: Test1234!", "409 Conflict 또는 차단", "완료",
         "test_registration_duplicate.py::test_registration_duplicate_email_blocked"],
        
        ["TC-025", "중복 방지", "대소문자 다른 중복 이메일", "API", "High", "API_Auto",
         "Mock 서버 실행 중", "1. Mixed case 이메일 등록\n2. Lowercase로 재시도\n3. 중복 확인",
         "email: CaseSensitive@Test.com", "대소문자 구분 확인", "완료",
         "test_registration_duplicate.py::test_registration_duplicate_email_case_insensitive"],
        
        ["TC-026", "UI", "UI - 정상 회원가입 플로우", "UI", "Critical", "UI_Auto",
         "웹 페이지 접속 가능", "1. 회원가입 페이지 접속\n2. 유효한 정보 입력\n3. 제출 버튼 클릭",
         "유효한 이메일과 비밀번호", "성공 메시지 표시", "완료",
         "test_registration_ui.py::test_ui_registration_happy_path"],
        
        ["TC-027", "UI", "UI - 잘못된 이메일 에러", "UI", "High", "UI_Auto",
         "웹 페이지 접속 가능", "1. 잘못된 이메일 입력\n2. 에러 메시지 확인",
         "invalid-email", "이메일 형식 에러 표시", "완료",
         "test_registration_ui.py::test_ui_invalid_email_format_error"],
        
        ["TC-028", "UI", "UI - 짧은 비밀번호 에러", "UI", "High", "UI_Auto",
         "웹 페이지 접속 가능", "1. 짧은 비밀번호 입력\n2. 제출\n3. 에러 확인",
         "password: short", "비밀번호 길이 에러", "완료",
         "test_registration_ui.py::test_ui_short_password_error"],
        
        ["TC-029", "UI", "UI - 중복 이메일 에러", "UI", "High", "UI_Auto",
         "웹 페이지 접속 가능", "1. 이미 등록된 이메일 입력\n2. 제출\n3. 에러 확인",
         "duplicate@test.com", "중복 이메일 에러", "완료",
         "test_registration_ui.py::test_ui_duplicate_email_error"],
        
        ["TC-030", "UI", "UI - 필수 필드 검증", "UI", "Medium", "UI_Auto",
         "웹 페이지 접속 가능", "1. 빈 폼 제출\n2. 필수 필드 검증",
         "empty fields", "필수 필드 에러", "완료",
         "test_registration_ui.py::test_ui_required_fields_validation"],
        
        ["TC-031", "UI", "UI - 실시간 입력 검증", "UI", "Low", "UI_Auto",
         "웹 페이지 접속 가능", "1. 잘못된 입력\n2. 포커스 이동\n3. 실시간 에러 확인",
         "various inputs", "실시간 검증 동작", "완료",
         "test_registration_ui.py::test_ui_realtime_validation_feedback"]
    ]
    
    for case in test_cases:
        ws_cases.append(case)
    
    # Style the header
    for cell in ws_cases[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Create Execution Tracking sheet
    ws_exec = wb.create_sheet("Execution Tracking")
    ws_exec.append(["TC_ID", "Last_Execution", "Last_Result", "Execution_Time", "Defects_Found", "Notes"])
    
    for i in range(1, 32):
        ws_exec.append([f"TC-{str(i).zfill(3)}", "Not Executed", "N/A", "N/A", 0, ""])
    
    # Style the header
    for cell in ws_exec[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    output_file = "test_cases.xlsx"
    wb.save(output_file)
    
    print(f"Test cases Excel file created: {output_file}")
    print(f"Total test cases: 31")
    print(f"- API Tests: 25")
    print(f"- UI Tests: 6")

if __name__ == "__main__":
    create_test_cases_excel()